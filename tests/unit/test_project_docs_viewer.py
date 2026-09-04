import io
import zipfile

import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from app.api import project_docs


def _write_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["name", "amount"])
    ws.append(["OHVIS", 5600])
    wb.save(path)


@pytest.mark.asyncio
async def test_project_docs_content_resolves_app_alias_and_converts_xlsx(tmp_path, monkeypatch):
    docs_dir = tmp_path / "docs"
    report_dir = docs_dir / "reports"
    report_dir.mkdir(parents=True)
    xlsx_path = report_dir / "sample.xlsx"
    _write_xlsx(xlsx_path)

    monkeypatch.setattr(
        project_docs,
        "SERVER_CONFIG",
        {
            "AADS": {
                "host": None,
                "paths": [{"base": "/app/docs", "label": "서버 문서"}],
            }
        },
    )
    monkeypatch.setattr(project_docs, "LOCAL_BASE_ALIASES", {"/app/docs": [str(docs_dir)]})

    response = await project_docs.get_doc_content(
        project="AADS",
        base_path="/app/docs",
        file_path="reports/sample.xlsx",
    )

    assert response["encoding"] == "text"
    assert response["mime_type"] == "text/csv"
    assert response["format"] == "excel-csv"
    assert response["converted_from"] == "xlsx"
    assert "## Sheet: Orders" in response["content"]
    assert "OHVIS,5600" in response["content"]


@pytest.mark.asyncio
async def test_project_docs_content_returns_downloadable_binary_for_archive(tmp_path, monkeypatch):
    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    zip_path = reports_dir / "bundle.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("readme.txt", "hello")

    monkeypatch.setattr(
        project_docs,
        "SERVER_CONFIG",
        {
            "AADS": {
                "host": None,
                "paths": [{"base": "/app/reports", "label": "서버 리포트"}],
            }
        },
    )
    monkeypatch.setattr(project_docs, "LOCAL_BASE_ALIASES", {"/app/reports": [str(reports_dir)]})

    response = await project_docs.get_doc_content(
        project="AADS",
        base_path="/app/reports",
        file_path="bundle.zip",
    )

    assert response["encoding"] == "base64"
    assert response["is_binary"] is True
    assert response["format"] == "binary"
    assert len(response["content"]) > 0


@pytest.mark.asyncio
async def test_project_docs_content_blocks_sensitive_relative_paths(tmp_path, monkeypatch):
    docs_dir = tmp_path / "docs"
    docs_dir.mkdir()
    (docs_dir / ".env").write_text("SECRET=1", encoding="utf-8")

    monkeypatch.setattr(
        project_docs,
        "SERVER_CONFIG",
        {
            "AADS": {
                "host": None,
                "paths": [{"base": "/app/docs", "label": "서버 문서"}],
            }
        },
    )
    monkeypatch.setattr(project_docs, "LOCAL_BASE_ALIASES", {"/app/docs": [str(docs_dir)]})

    with pytest.raises(HTTPException) as excinfo:
        await project_docs.get_doc_content(
            project="AADS",
            base_path="/app/docs",
            file_path=".env",
        )

    assert excinfo.value.status_code == 400


@pytest.mark.asyncio
async def test_project_docs_content_repairs_legacy_aads_go100_route(monkeypatch):
    async def fake_run_cmd(cmd, timeout=10):
        remote_cmd = cmd[-1]
        if "test -f" in remote_cmd and "/root/kis-autotrade-v4/docs/reports/GO100-303.md" in remote_cmd:
            return "exists"
        if remote_cmd == "cat /root/kis-autotrade-v4/docs/reports/GO100-303.md":
            return "# GO100 report"
        return ""

    monkeypatch.setattr(project_docs, "_run_cmd", fake_run_cmd)

    response = await project_docs.get_doc_content(
        project="AADS",
        base_path="/app/docs",
        file_path="reports/GO100-303.md",
    )

    assert response["project"] == "GO100"
    assert response["file_path"] == "reports/GO100-303.md"
    assert response["full_path"] == "/root/kis-autotrade-v4/docs/reports/GO100-303.md"
    assert response["content"] == "# GO100 report"


@pytest.mark.asyncio
async def test_project_docs_content_falls_back_from_go100_reports_to_docs_reports(monkeypatch):
    async def fake_run_cmd(cmd, timeout=10):
        remote_cmd = cmd[-1]
        if "test -f" in remote_cmd and "/root/kis-autotrade-v4/docs/reports/GO100-303.md" in remote_cmd:
            return "exists"
        if remote_cmd == "cat /root/kis-autotrade-v4/docs/reports/GO100-303.md":
            return "# GO100 docs report"
        return ""

    monkeypatch.setattr(project_docs, "_run_cmd", fake_run_cmd)

    response = await project_docs.get_doc_content(
        project="GO100",
        base_path="/root/kis-autotrade-v4/reports",
        file_path="GO100-303.md",
    )

    assert response["project"] == "GO100"
    assert response["file_path"] == "GO100-303.md"
    assert response["full_path"] == "/root/kis-autotrade-v4/docs/reports/GO100-303.md"
    assert response["content"] == "# GO100 docs report"


def test_excel_bytes_to_csv_text_uses_all_sheets():
    wb = Workbook()
    ws = wb.active
    ws.title = "First"
    ws.append(["a", "b"])
    second = wb.create_sheet("Second")
    second.append(["x", "y"])
    stream = io.BytesIO()
    wb.save(stream)

    content = project_docs._excel_bytes_to_csv_text(stream.getvalue(), "multi.xlsx")

    assert "# multi.xlsx CSV preview" in content
    assert "## Sheet: First" in content
    assert "a,b" in content
    assert "## Sheet: Second" in content
    assert "x,y" in content


@pytest.mark.asyncio
async def test_project_docs_scan_include_filter_matches_relative_path(tmp_path, monkeypatch):
    docs_dir = tmp_path / "docs"
    nested = docs_dir / "go100" / "user-guide"
    nested.mkdir(parents=True)
    (nested / "onboarding.md").write_text("# onboarding", encoding="utf-8")
    (docs_dir / "README.md").write_text("# generic", encoding="utf-8")

    results = await project_docs._scan_local(str(docs_dir), include=["go100/"])

    assert [item["path"] for item in results] == ["go100/user-guide/onboarding.md"]


@pytest.mark.asyncio
async def test_go100_document_status_scans_api_and_artifacts_paths(monkeypatch):
    captured_bases = []

    async def fake_scan_remote(host, base, exclude=None, include=None):
        captured_bases.append((host, base, tuple(include or ()), tuple(exclude or ())))
        return []

    monkeypatch.setattr(project_docs, "_scan_remote", fake_scan_remote)

    await project_docs._scan_project("GO100", project_docs.SERVER_CONFIG["GO100"])

    bases = {base for _, base, _, _ in captured_bases}
    assert "/root/kis-autotrade-v4/docs/api" in bases
    assert "/root/kis-autotrade-v4/docs/plans" in bases
    assert "/root/kis-autotrade-v4/docs/handover" in bases
    assert "/root/kis-autotrade-v4/artifacts/go100" in bases

    catch_all = next(
        item for item in captured_bases
        if item[1] == "/root/kis-autotrade-v4/docs"
    )
    assert "api/" in catch_all[3]
    assert "kis-api-portal/" in catch_all[3]
