"""Browser collectors for Yeoljeong delivery-platform ledgers.

The collectors deliberately stop at CAPTCHA, OTP, device verification, or
terms-consent screens.  They never persist credentials, browser storage, raw
HTML, or downloaded source files.  Portal selectors are kept in this module so
DOM changes do not leak into the finance service or API layer.
"""
from __future__ import annotations

import csv
import hashlib
import os
import re
import shutil
import tempfile
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any


PORTAL_CONFIG: dict[str, dict[str, Any]] = {
    "baemin": {
        "label": "배민셀프서비스",
        "login_url": "https://biz-member.baemin.com/login?returnUrl=https%3A%2F%2Fself.baemin.com%2F",
        "sections": {
            "sales": ("매출", "주문내역", "주문 관리"),
            "settlements": ("정산", "정산내역", "부가세 신고자료"),
            "reviews": ("리뷰", "리뷰 관리"),
        },
    },
    "coupangeats": {
        "label": "쿠팡이츠 사장님",
        "login_url": "https://store.coupangeats.com/merchant/login",
        "sections": {
            "sales": ("매출", "주문내역", "주문 관리"),
            "settlements": ("정산", "정산 관리"),
            "reviews": ("리뷰", "리뷰 관리"),
        },
    },
    "yogiyo": {
        "label": "요기요 사장님",
        "login_url": "https://ceo.yogiyo.co.kr/login",
        "sections": {
            "sales": ("매출", "주문내역", "주문 관리"),
            "settlements": ("정산", "정산내역"),
            "reviews": ("리뷰", "리뷰 관리"),
        },
    },
    "ddangyo": {
        "label": "땡겨요 사장님",
        "login_url": "https://boss.ddangyo.com/",
        "dismiss_selectors": (
            "#mf_wfm_side_SMWCO000002P04_close",
            "input[type='button'][value='닫기']",
        ),
        "sections": {
            "sales": ("주문내역",),
            "settlements": ("정산내역",),
            "reviews": ("리뷰관리",),
        },
    },
}

CHALLENGE_TERMS = (
    "captcha",
    "캡차",
    "보안문자",
    "자동입력방지",
    "2차 인증",
    "추가 인증",
    "본인인증",
    "휴대폰 인증",
    "기기 인증",
    "인증번호",
    "약관 동의",
)
DDANGYO_NUMERIC_CAPTCHA_TERMS = ("captcha", "캡차", "보안문자", "자동입력방지", "숫자를 입력")
LOGIN_FAILURE_TERMS = ("비밀번호가 일치", "아이디 또는 비밀번호", "로그인 정보를 확인", "로그인 실패")
LOGGED_OUT_TERMS = ("로그인해주세요", "로그인이 필요", "사장님 로그인")
EXPORT_TERMS = ("엑셀", "excel", "csv", "다운로드", "내보내기")
SEARCH_TERMS = ("조회", "검색", "적용")
SECURITY_BLOCK_TERMS = (
    "보안 위배 접근 제한",
    "올바르지 않은 요청",
    "access denied",
    "forbidden",
)
LOGIN_SELECTOR_CONFIG: dict[str, dict[str, tuple[str, ...]]] = {
    "baemin": {
        "username": (
            "input[autocomplete='username']",
            "input[name='id']",
            "input[name*='id' i]",
            "input[name*='user' i]",
            "input[type='email']",
            "input[type='text']",
        ),
        "password": (
            "input[autocomplete='current-password']",
            "input[name='password']",
            "input[name*='pw' i]",
            "input[type='password']",
        ),
        "submit": ("button[type='submit']", "input[type='submit']"),
    },
    "coupangeats": {
        "username": (
            "input[autocomplete='username']",
            "input[name*='email' i]",
            "input[name*='id' i]",
            "input[name*='user' i]",
            "input[placeholder*='아이디']",
            "input[placeholder*='이메일']",
            "input[type='email']",
            "input[type='text']",
        ),
        "password": (
            "input[autocomplete='current-password']",
            "input[name*='password' i]",
            "input[name*='pw' i]",
            "input[placeholder*='비밀번호']",
            "input[type='password']",
        ),
        "submit": ("button[type='submit']", "input[type='submit']"),
    },
    "yogiyo": {
        "username": (
            "input[autocomplete='username']",
            "input[name*='id' i]",
            "input[name*='email' i]",
            "input[name*='user' i]",
            "input[placeholder*='아이디']",
            "input[placeholder*='이메일']",
            "input[type='email']",
            "input[type='text']",
        ),
        "password": (
            "input[autocomplete='current-password']",
            "input[name*='password' i]",
            "input[name*='pw' i]",
            "input[placeholder*='비밀번호']",
            "input[type='password']",
        ),
        "submit": ("button[type='submit']", "input[type='submit']"),
    },
    "ddangyo": {
        "username": (
            "#mf_wfm_login_id",
            "#mf_ipt_usrId",
            "#userId",
            "input[name*='user' i]",
            "input[name*='id' i]",
            "input[placeholder*='아이디']",
            "input[type='text']",
        ),
        "password": (
            "#mf_wfm_login_pw",
            "#mf_ipt_pw",
            "#password",
            "input[name*='password' i]",
            "input[name*='pw' i]",
            "input[placeholder*='비밀번호']",
            "input[type='password']",
        ),
        "submit": (
            "#mf_btn_webLogin",
            "input[type='button'][value*='로그인']",
            "button[type='submit']",
            "input[type='submit']",
        ),
    },
}
STORAGE_STATE_ACCOUNT_KEYS = (
    "storage_state_path",
    "browser_storage_state_path",
    "baemin_storage_state_path",
)
STORAGE_STATE_ENV_KEYS = (
    "YEOLJEONG_BAEMIN_STORAGE_STATE",
    "BAEMIN_STORAGE_STATE_PATH",
    "AADS_BROWSER_BRIDGE_STORAGE_STATE",
)


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _number(value: Any) -> int:
    text = re.sub(r"[^0-9.-]", "", _clean(value))
    if not text or text in {"-", ".", "-."}:
        return 0
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


def _date(value: Any) -> str:
    text = _clean(value)
    match = re.search(r"(20\d{2})[./-](\d{1,2})[./-](\d{1,2})", text)
    if not match:
        return ""
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
    except ValueError:
        return ""


def _first(row: dict[str, Any], terms: tuple[str, ...]) -> str:
    for term in terms:
        lowered_term = term.lower()
        for key, value in row.items():
            if lowered_term in _clean(key).lower():
                cleaned = _clean(value)
                if cleaned:
                    return cleaned
    return ""


def _source_id(service: str, kind: str, row: dict[str, Any]) -> str:
    explicit = _first(row, ("주문번호", "정산번호", "리뷰번호", "order id", "id"))
    material = explicit or "|".join(f"{_clean(k)}={_clean(v)}" for k, v in sorted(row.items(), key=lambda item: str(item[0])))
    return hashlib.sha256(f"{service}|{kind}|{material}".encode("utf-8")).hexdigest()[:32]


def normalize_record(
    service: str,
    kind: str,
    row: dict[str, Any],
    business_id: str,
    branch: str,
) -> dict[str, Any]:
    """Convert a portal row to the minimal canonical ledger schema."""
    source_id = _source_id(service, kind, row)
    occurred_on = _date(
        _first(row, ("거래일", "주문일", "매출일", "정산일", "입금(예정)일", "입금일", "리뷰일", "작성일", "date", "일자"))
    )
    record: dict[str, Any] = {
        "id": hashlib.sha256(f"{business_id}|{branch}|{service}|{kind}|{source_id}".encode("utf-8")).hexdigest(),
        "source_id": source_id,
        "business_id": business_id,
        "branch": branch,
        "service": service,
        "platform": service,
        "record_type": kind,
        "occurred_on": occurred_on,
        "collected_at": datetime.now().astimezone().isoformat(timespec="seconds"),
    }
    if kind == "sales":
        record.update(
            {
                "order_id": _first(row, ("주문번호", "order id", "주문 id")),
                "gross_amount": _number(_first(row, ("총매출", "주문금액", "결제금액", "매출액", "판매금액"))),
                "discount_amount": _number(_first(row, ("할인", "쿠폰"))),
                "delivery_fee": _number(_first(row, ("배달팁", "배달비"))),
                "order_status": _first(row, ("주문상태", "상태")),
            }
        )
    elif kind == "settlements":
        record.update(
            {
                "settlement_id": _first(row, ("정산번호", "지급번호", "id")),
                "sales_amount": _number(_first(row, ("매출액", "주문금액", "판매금액"))),
                "fee_amount": _number(_first(row, ("중개수수료", "수수료"))),
                "vat_amount": _number(_first(row, ("부가세", "vat"))),
                "settlement_amount": _number(_first(row, ("정산금액", "지급금액", "입금(예정)금액", "입금액"))),
                "settlement_status": _first(row, ("정산상태", "지급상태", "입금상태", "상태")),
            }
        )
    else:
        record.update(
            {
                "review_id": _first(row, ("리뷰번호", "review id", "id")) or source_id,
                "rating": _number(_first(row, ("평점", "별점", "rating"))),
                "review_text": _first(row, ("리뷰내용", "리뷰", "내용", "후기"))[:4000],
                "reply_status": _first(row, ("답글상태", "답변상태", "사장님 답글")),
            }
        )
    return record


def _read_download(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [_clean(value) or f"column_{index + 1}" for index, value in enumerate(values[0])]
        return [dict(zip(headers, row)) for row in values[1:] if any(value not in (None, "") for value in row)]
    raw = path.read_bytes()
    text = ""
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        return []
    return [dict(row) for row in csv.DictReader(text.splitlines())]


def _login_selectors(service: str) -> dict[str, tuple[str, ...]]:
    fallback = {
        "username": (
            "input[autocomplete='username']",
            "input[name*='id' i]",
            "input[name*='user' i]",
            "input[type='email']",
            "input[type='text']",
        ),
        "password": ("input[autocomplete='current-password']", "input[type='password']"),
        "submit": (
            "button[type='submit']",
            "input[type='submit']",
            "input[type='button'][value*='로그인']",
        ),
    }
    configured = LOGIN_SELECTOR_CONFIG.get(str(service or "").strip().lower(), {})
    return {
        "username": tuple(configured.get("username") or fallback["username"]),
        "password": tuple(configured.get("password") or fallback["password"]),
        "submit": tuple(configured.get("submit") or fallback["submit"]),
    }


def _fill_login_dom(page: Any, service: str, username: str, password: str) -> dict[str, Any]:
    selectors = _login_selectors(service)
    try:
        result = page.evaluate(
            r"""
            ({username, password, usernameSelectors, passwordSelectors, submitSelectors}) => {
              const visible = element => {
                if (!element) return false;
                const style = window.getComputedStyle(element);
                const rect = element.getBoundingClientRect();
                return style.visibility !== 'hidden'
                  && style.display !== 'none'
                  && rect.width > 0
                  && rect.height > 0
                  && element.type !== 'hidden'
                  && !element.disabled;
              };
              const firstVisible = selectors => {
                for (const selector of selectors) {
                  try {
                    const found = [...document.querySelectorAll(selector)].find(visible);
                    if (found) return found;
                  } catch (_) {}
                }
                return null;
              };
              const setNativeValue = (element, value) => {
                const descriptor = Object.getOwnPropertyDescriptor(Object.getPrototypeOf(element), 'value');
                if (descriptor && descriptor.set) descriptor.set.call(element, value);
                else element.value = value;
                element.dispatchEvent(new Event('input', {bubbles: true}));
                element.dispatchEvent(new Event('change', {bubbles: true}));
                element.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true, key: 'Unidentified'}));
              };
              const userInput = firstVisible(usernameSelectors);
              const passwordInput = firstVisible(passwordSelectors);
              if (!userInput || !passwordInput) {
                return {filled: false, clicked: false, reason: 'LOGIN_FORM_NOT_FOUND'};
              }
              userInput.focus();
              setNativeValue(userInput, username);
              passwordInput.focus();
              setNativeValue(passwordInput, password);
              let submit = firstVisible(submitSelectors);
              if (!submit) {
                const labels = ['로그인', 'login', 'sign in', '확인'];
                submit = [...document.querySelectorAll('button,a,[role="button"],input[type="button"],input[type="submit"]')]
                  .find(element => {
                    if (!visible(element)) return false;
                    const text = String(element.innerText || element.textContent || element.value || '').trim().toLowerCase();
                    return labels.some(label => text.includes(label));
                  });
              }
              if (submit) {
                submit.click();
                return {filled: true, clicked: true, reason: ''};
              }
              const form = passwordInput.closest('form') || userInput.closest('form');
              if (form) {
                if (typeof form.requestSubmit === 'function') form.requestSubmit();
                else form.submit();
                return {filled: true, clicked: true, reason: 'FORM_SUBMIT'};
              }
              passwordInput.dispatchEvent(new KeyboardEvent('keydown', {bubbles: true, key: 'Enter', code: 'Enter'}));
              passwordInput.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true, key: 'Enter', code: 'Enter'}));
              return {filled: true, clicked: false, reason: 'ENTER_DISPATCHED'};
            }
            """,
            {
                "username": username,
                "password": password,
                "usernameSelectors": list(selectors["username"]),
                "passwordSelectors": list(selectors["password"]),
                "submitSelectors": list(selectors["submit"]),
            },
        )
        return result if isinstance(result, dict) else {"filled": bool(result), "clicked": False, "reason": ""}
    except Exception as exc:
        return {"filled": False, "clicked": False, "reason": exc.__class__.__name__}


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[str]]] = []
        self.text: list[str] = []
        self._table_index = -1
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style"}:
            self._skip_depth += 1
            return
        if self._skip_depth:
            return
        if tag == "table":
            self.tables.append([])
            self._table_index = len(self.tables) - 1
        elif tag == "tr" and self._table_index >= 0:
            self._row = []
        elif tag in {"th", "td"} and self._row is not None:
            self._cell = []
        elif tag in {"br", "p", "div", "li"}:
            self.text.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style"} and self._skip_depth:
            self._skip_depth -= 1
            return
        if self._skip_depth:
            return
        if tag in {"th", "td"} and self._row is not None and self._cell is not None:
            self._row.append(_clean(" ".join(self._cell)))
            self._cell = None
        elif tag == "tr" and self._table_index >= 0 and self._row is not None:
            if any(self._row):
                self.tables[self._table_index].append(self._row)
            self._row = None
        elif tag == "table":
            self._table_index = -1
        elif tag in {"p", "div", "li"}:
            self.text.append("\n")

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        cleaned = _clean(data)
        if not cleaned:
            return
        if self._cell is not None:
            self._cell.append(cleaned)
        self.text.append(cleaned)


def _rows_to_dicts(rows: list[list[str]]) -> list[dict[str, Any]]:
    compact = [[_clean(cell) for cell in row] for row in rows if any(_clean(cell) for cell in row)]
    if len(compact) < 2:
        return []
    header_index = 0
    for index, row in enumerate(compact[:5]):
        joined = " ".join(row)
        if any(term in joined for term in ("일", "번호", "금액", "상태", "리뷰", "별점", "평점")):
            header_index = index
            break
    headers = [value or f"column_{index + 1}" for index, value in enumerate(compact[header_index])]
    parsed: list[dict[str, Any]] = []
    for row in compact[header_index + 1 :]:
        if len(row) < len(headers):
            row = [*row, *([""] * (len(headers) - len(row)))]
        parsed.append({headers[index] if index < len(headers) else f"column_{index + 1}": value for index, value in enumerate(row)})
    return parsed


def _parse_delimited_text(text: str) -> list[dict[str, Any]]:
    lines = [line for line in text.splitlines() if _clean(line)]
    if len(lines) < 2:
        return []
    sample = "\n".join(lines[:20])
    delimiters = ["\t", ",", ";", "|"]
    delimiter = "\t" if "\t" in sample else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="".join(delimiters))
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    reader = csv.DictReader(lines, delimiter=delimiter)
    if not reader.fieldnames:
        return []
    return [dict(row) for row in reader if any(_clean(value) for value in row.values())]


def parse_portal_export(
    service: str,
    kind: str,
    source_text: str,
    business_id: str,
    branch: str,
) -> dict[str, Any]:
    """Parse copied/saved portal tables from an already-authenticated PC browser.

    This path is intentionally offline: it accepts only user-provided page/table
    content and does not attempt to bypass portal IP, CAPTCHA, OTP, or device
    checks from the server.
    """
    normalized_service = str(service or "").strip()
    normalized_kind = str(kind or "").strip()
    if normalized_service not in PORTAL_CONFIG:
        return {"status": "failed", "error_code": "UNSUPPORTED_PLATFORM", "records": {}, "diagnostics": {}}
    if normalized_kind not in {"sales", "settlements", "reviews"}:
        return {"status": "failed", "error_code": "UNSUPPORTED_RECORD_TYPE", "records": {}, "diagnostics": {}}

    raw = str(source_text or "")
    if not raw.strip():
        return {"status": "failed", "error_code": "EMPTY_SOURCE", "records": {}, "diagnostics": {}}

    table_rows: list[dict[str, Any]] = []
    source = "delimited_text"
    if "<" in raw and ">" in raw:
        parser = _HtmlTableParser()
        parser.feed(raw)
        for table in parser.tables:
            table_rows.extend(_rows_to_dicts(table))
        source = "html_table" if table_rows else "html_text"
        if not table_rows:
            table_rows = _parse_delimited_text("\n".join(parser.text))
    if not table_rows:
        table_rows = _parse_delimited_text(raw)
    if not table_rows:
        return {
            "status": "partial",
            "error_code": "NO_PARSEABLE_ROWS",
            "records": {normalized_kind: []},
            "diagnostics": {"source": source},
            "message": "복사/저장한 페이지에서 표 형식 데이터를 찾지 못했습니다.",
        }

    normalized = [
        normalize_record(normalized_service, normalized_kind, row, str(business_id or ""), str(branch or ""))
        for row in table_rows
    ]
    return {
        "status": "succeeded",
        "error_code": "",
        "records": {normalized_kind: normalized},
        "diagnostics": {"source": source, "source_rows": len(table_rows)},
    }


def _click_first(page: Any, labels: tuple[str, ...], timeout: int = 2500) -> bool:
    for label in labels:
        for exact in (True, False):
            pattern = re.compile(re.escape(label), re.I)
            for role in ("button", "link", None):
                try:
                    matches = page.get_by_role(role, name=pattern, exact=exact) if role else page.get_by_text(pattern, exact=exact)
                except Exception:
                    continue
                for index in range(min(matches.count(), 20)):
                    locator = matches.nth(index)
                    try:
                        if locator.is_visible(timeout=500):
                            if role == "link":
                                locator.evaluate("element => element.click()")
                            else:
                                try:
                                    locator.click(timeout=timeout)
                                except Exception:
                                    locator.click(timeout=timeout, force=True)
                            page.wait_for_timeout(800)
                            return True
                    except Exception:
                        continue
    return False


def _fill_login(page: Any, username: str, password: str, service: str = "") -> bool:
    # Several merchant portals render their login controls after
    # ``domcontentloaded``.  Wait for the dynamic form instead of treating the
    # first empty DOM snapshot as a missing/changed login page.
    try:
        page.wait_for_load_state("networkidle", timeout=8000)
    except Exception:
        pass
    try:
        page.wait_for_selector("input[type='password']", state="visible", timeout=8000)
    except Exception:
        pass

    selectors = _login_selectors(service)
    def first_visible(selectors: tuple[str, ...]) -> Any | None:
        for selector in selectors:
            locator = page.locator(selector).first
            try:
                if locator.count() and locator.is_visible(timeout=500):
                    return locator
            except Exception:
                continue
        return None

    user_input = first_visible(selectors["username"])
    password_input = first_visible(selectors["password"])
    clicked = False
    if user_input is None or password_input is None:
        dom_result = _fill_login_dom(page, service, username, password)
        if not dom_result.get("filled"):
            return False
        clicked = bool(dom_result.get("clicked"))
    else:
        user_input.fill(username)
        password_input.fill(password)
        clicked = _click_first(page, ("로그인", "login"), timeout=4000)
    if not clicked:
        clicked = False
        for selector in selectors["submit"]:
            control = page.locator(selector).first
            try:
                if control.count() and control.is_visible(timeout=500):
                    control.click(timeout=4000)
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked and password_input is not None:
            password_input.press("Enter")
    page.wait_for_timeout(5000)
    try:
        page.wait_for_load_state("networkidle", timeout=5000)
    except Exception:
        pass
    return True


def _page_state(page: Any, service: str = "") -> tuple[str, str]:
    body = _clean(page.locator("body").inner_text(timeout=5000)).lower()
    if any(term in body for term in SECURITY_BLOCK_TERMS):
        return "portal_action_required", "BAEMIN_SECURITY_BLOCKED"
    if service == "ddangyo" and any(term in body for term in DDANGYO_NUMERIC_CAPTCHA_TERMS):
        return "portal_action_required", "DDANGYO_NUMERIC_CAPTCHA_REQUIRED"
    if any(term in body for term in CHALLENGE_TERMS):
        return "portal_action_required", "PORTAL_AUTH_CHALLENGE"
    if any(term in body for term in LOGIN_FAILURE_TERMS):
        return "failed", "PORTAL_CREDENTIAL_REJECTED"
    if any(term in body for term in LOGGED_OUT_TERMS):
        return "failed", "PORTAL_LOGIN_NOT_COMPLETED"
    password_inputs = page.locator("input[type='password']")
    for index in range(password_inputs.count()):
        try:
            if password_inputs.nth(index).is_visible(timeout=500):
                return "failed", "PORTAL_LOGIN_NOT_COMPLETED"
        except Exception:
            continue
    return "authenticated", ""


def _security_block_result(page: Any, response: Any | None) -> dict[str, Any] | None:
    status = int(getattr(response, "status", 0) or 0)
    try:
        body = _clean(page.locator("body").inner_text(timeout=3000)).lower()
    except Exception:
        body = ""
    if status in {401, 403} or any(term in body for term in SECURITY_BLOCK_TERMS):
        return {
            "status": "portal_action_required",
            "error_code": "BAEMIN_SECURITY_BLOCKED",
            "records": {},
            "diagnostics": {"http_status": status or ""},
            "message": "배민 포털이 서버 자동접속을 보안 정책으로 차단했습니다. PC 인증 세션 또는 정산 CSV 업로드가 필요합니다.",
        }
    return None


def _login_url(account: dict[str, Any], config: dict[str, Any]) -> str:
    service = str(account.get("service") or "")
    configured = str(account.get("login_url") or "").strip()
    if service == "baemin" and "biz-member.baemin.com/login" not in configured:
        return str(config["login_url"])
    return configured or str(config["login_url"])


def _portal_home_url(account: dict[str, Any], config: dict[str, Any]) -> str:
    service = str(account.get("service") or "")
    configured = str(account.get("portal_home_url") or account.get("home_url") or "").strip()
    if configured:
        return configured
    if service == "baemin":
        return "https://self.baemin.com/"
    return _login_url(account, config)


def _storage_state_path(account: dict[str, Any]) -> str:
    for key in STORAGE_STATE_ACCOUNT_KEYS:
        value = str(account.get(key) or "").strip()
        if value and Path(value).expanduser().is_file():
            return str(Path(value).expanduser())
    for key in STORAGE_STATE_ENV_KEYS:
        value = str(os.environ.get(key) or "").strip()
        if value and Path(value).expanduser().is_file():
            return str(Path(value).expanduser())
    return ""


def _set_period(page: Any, date_from: str, date_to: str) -> None:
    date_inputs = page.locator("input[type='date']")
    if date_inputs.count() >= 1:
        date_inputs.nth(0).fill(date_from)
    if date_inputs.count() >= 2:
        date_inputs.nth(1).fill(date_to)
    if date_inputs.count() < 2:
        start_inputs = page.locator("input[title*='시작 날짜']")
        end_inputs = page.locator("input[title*='종료 날짜']")
        if start_inputs.count():
            start_inputs.first.fill(date_from)
        if end_inputs.count():
            end_inputs.first.fill(date_to)
    _click_first(page, SEARCH_TERMS)


def _dismiss_optional_prompts(page: Any, config: dict[str, Any]) -> None:
    """Close optional marketing/notice prompts without accepting new terms."""
    for selector in config.get("dismiss_selectors", ()):
        try:
            locator = page.locator(selector).first
            if locator.count() and locator.is_visible(timeout=500):
                locator.click(timeout=2500, force=True)
                page.wait_for_timeout(500)
        except Exception:
            continue


def _scrape_table(page: Any) -> list[dict[str, Any]]:
    return page.locator("table").first.evaluate(
        """
        table => {
          const headers = [...table.querySelectorAll('thead th')].map((x, i) => x.innerText.trim() || `column_${i + 1}`);
          const rows = [...table.querySelectorAll('tbody tr')];
          return rows.map(row => {
            const cells = [...row.querySelectorAll('th,td')].map(x => x.innerText.trim());
            return Object.fromEntries(cells.map((value, i) => [headers[i] || `column_${i + 1}`, value]));
          });
        }
        """
    )


def _scrape_list_table(page: Any) -> list[dict[str, Any]]:
    """Scrape WebSquare-style list tables used by the Ddangyo portal."""
    return page.locator("section.C_table").first.evaluate(
        """
        section => {
          const header = section.querySelector('ul.C_table_list.head');
          if (!header) return [];
          const headers = [...header.children].map((cell, index) => cell.innerText.trim() || `column_${index + 1}`);
          return [...section.querySelectorAll('ul.C_table_list.body')].map(row => {
            const cells = [...row.children].map(cell => cell.innerText.trim());
            return Object.fromEntries(cells.map((value, index) => [headers[index] || `column_${index + 1}`, value]));
          });
        }
        """
    )


def _scrape_review_cards(page: Any) -> list[dict[str, Any]]:
    """Scrape Ddangyo review cards without persisting owner replies or raw HTML."""
    return page.locator("#mf_wfm_contents_wfm_tabFrame_gen_generator1 > section.C_data_box").evaluate_all(
        r"""
        cards => cards.map(card => {
          const text = card.innerText.trim();
          const lines = text.split(/\n+/).map(value => value.trim()).filter(Boolean);
          const dateMatch = text.match(/20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}/);
          const review = card.querySelector('.re_cont_txt');
          return {
            '작성자': lines[0] || '',
            '작성일': dateMatch ? dateMatch[0] : '',
            '리뷰내용': review ? review.innerText.trim() : '',
            '음식평가': lines.includes('맛있어요') ? '맛있어요' : '',
            '답글상태': /사장님20\d{2}-\d{2}-\d{2}/.test(text) ? '답변완료' : '미답변',
          };
        }).filter(row => row['작성일'] && row['리뷰내용'])
        """
    )


def _collect_section(
    page: Any,
    labels: tuple[str, ...],
    date_from: str,
    date_to: str,
    download_dir: Path,
    config: dict[str, Any],
    kind: str,
) -> tuple[list[dict[str, Any]], str]:
    if not _click_first(page, labels):
        return [], "SECTION_NOT_FOUND"
    _dismiss_optional_prompts(page, config)
    _set_period(page, date_from, date_to)
    for export_label in EXPORT_TERMS:
        try:
            button = page.get_by_text(re.compile(re.escape(export_label), re.I), exact=False).first
            if not button.count() or not button.is_visible(timeout=400):
                continue
            with page.expect_download(timeout=5000) as download_info:
                button.click(timeout=2500)
            download = download_info.value
            target = download_dir / (download.suggested_filename or "export.csv")
            download.save_as(target)
            rows = _read_download(target)
            target.unlink(missing_ok=True)
            return rows, "download"
        except Exception:
            continue
    try:
        if kind == "reviews" and page.locator("#mf_wfm_contents_wfm_tabFrame_gen_generator1 > section.C_data_box").count():
            return _scrape_review_cards(page), "review_cards"
        if page.locator("section.C_table ul.C_table_list.body").count():
            return _scrape_list_table(page), "list_table"
        if page.locator("table").count():
            return _scrape_table(page), "table"
    except Exception:
        pass
    return [], "NO_EXPORT_OR_TABLE"


def collect_account(
    account: dict[str, Any],
    password: str,
    date_from: str,
    date_to: str,
) -> dict[str, Any]:
    """Log in and collect sales, settlement, and review rows for one account."""
    service = str(account.get("service") or "")
    config = PORTAL_CONFIG.get(service)
    if not config:
        return {"status": "failed", "error_code": "UNSUPPORTED_PLATFORM", "records": {}}
    storage_state = _storage_state_path(account)
    if not password and not storage_state:
        return {"status": "credential_required", "error_code": "CREDENTIAL_REQUIRED", "records": {}}
    temp_dir = Path(tempfile.mkdtemp(prefix=f"yf-{service}-"))
    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            context_kwargs: dict[str, Any] = {"accept_downloads": True, "locale": "ko-KR"}
            if storage_state:
                context_kwargs["storage_state"] = storage_state
            context = browser.new_context(**context_kwargs)
            page = context.new_page()
            auth_mode = "storage_state" if storage_state else "password_login"
            response = page.goto(
                _portal_home_url(account, config) if storage_state else _login_url(account, config),
                wait_until="domcontentloaded",
                timeout=30000,
            )
            blocked = _security_block_result(page, response)
            if blocked:
                blocked.setdefault("diagnostics", {})["auth_mode"] = auth_mode
                browser.close()
                return blocked
            state, error_code = _page_state(page, service)
            if state != "authenticated":
                if not password:
                    browser.close()
                    return {
                        "status": "credential_required",
                        "error_code": "STORAGE_STATE_EXPIRED",
                        "records": {},
                        "diagnostics": {"auth_mode": auth_mode},
                        "message": "저장된 PC 브라우저 인증 세션이 만료됐거나 로그인 상태가 아닙니다.",
                    }
                response = page.goto(_login_url(account, config), wait_until="domcontentloaded", timeout=30000)
                blocked = _security_block_result(page, response)
                if blocked:
                    blocked.setdefault("diagnostics", {})["auth_mode"] = auth_mode
                    browser.close()
                    return blocked
                if not _fill_login(page, str(account.get("username") or ""), password, service):
                    browser.close()
                    return {
                        "status": "portal_action_required",
                        "error_code": "LOGIN_FORM_NOT_FOUND",
                        "records": {},
                        "diagnostics": {"auth_mode": auth_mode},
                    }
                state, error_code = _page_state(page, service)
            if state != "authenticated":
                browser.close()
                return {"status": state, "error_code": error_code, "records": {}, "diagnostics": {"auth_mode": auth_mode}}
            _dismiss_optional_prompts(page, config)

            collected: dict[str, list[dict[str, Any]]] = {}
            diagnostics: dict[str, str] = {"auth_mode": auth_mode}
            for kind, labels in config["sections"].items():
                rows, source = _collect_section(page, labels, date_from, date_to, temp_dir, config, kind)
                collected[kind] = [
                    normalize_record(service, kind, row, str(account.get("business_id") or ""), str(account.get("branch") or ""))
                    for row in rows
                ]
                diagnostics[kind] = source
            browser.close()
            total = sum(len(rows) for rows in collected.values())
            return {
                "status": "succeeded" if total else "partial",
                "error_code": "" if total else "AUTHENTICATED_NO_ROWS",
                "records": collected,
                "diagnostics": diagnostics,
            }
    except Exception as exc:
        error_name = type(exc).__name__.upper()
        return {"status": "failed", "error_code": f"COLLECTOR_{error_name}", "records": {}}
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
