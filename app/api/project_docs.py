"""
프로젝트별 문서 통합 조회 API.
프로젝트 서버의 docs, reports 디렉토리를 스캔하여
프로젝트별로 분류된 문서 목록과 내용을 제공한다.
"""
from __future__ import annotations

import asyncio
import base64
import csv
import html
import io
import json
import mimetypes
import os
import re
import shlex
import time
import zipfile
from pathlib import Path
from typing import Optional

import structlog
from fastapi import APIRouter, HTTPException, Query

router = APIRouter()
logger = structlog.get_logger()

# ── 캐시 ──
_cache: dict = {"data": None, "ts": 0}
CACHE_TTL = 300  # 5분
PERSISTENT_CACHE_FILE = Path(os.getenv("PROJECT_DOCS_CACHE_FILE", "/tmp/aads_project_docs_cache.json"))

# ── 서버/프로젝트 경로 매핑 ──
SERVER_CONFIG = {
    "AADS": {
        "host": None,  # 로컬
        "paths": [
            {"base": "/app/docs", "label": "서버 문서"},
            {"base": "/app/reports", "label": "서버 리포트"},
            {"base": "/root/aads/aads-server/docs", "label": "서버 문서"},
            {"base": "/root/aads/aads-server/reports", "label": "서버 리포트"},
            {"base": "/root/aads/aads-docs/docs", "label": "공용 문서"},
            {"base": "/root/aads/aads-docs/reports", "label": "공용 리포트", "exclude": ["ceo-documents/_index.json"]},
            {"base": "/root/aads/aads-dashboard/docs", "label": "대시보드 문서"},
            {"base": "/root/aads/aads-dashboard/reports", "label": "대시보드 리포트"},
            {"base": "/root/aads/aads-core/docs", "label": "코어 문서"},
            {"base": "/root/aads/aads-core/reports", "label": "코어 리포트"},
            {"base": "/app/app/static/docs", "label": "정적 문서"},
            {"base": "/app/app/static/reports", "label": "정적 리포트"},
            {"base": "/app/app/static/preview", "label": "프리뷰"},
            {"base": "/app/app/static/gallery", "label": "갤러리"},
            {"base": "/root/aads/aads-dashboard/public/reports", "label": "대시보드 공개 리포트"},
            {"base": "/root/aads/aads-dashboard/public/exports", "label": "대시보드 내보내기"},
        ],
    },
    "KIS": {
        "host": "contabo14",
        "paths": [
            {"base": "/root/kis-autotrade-v4/docs", "label": "문서",
             "exclude": ["kis-api-portal", "GO100", "go100"]},
        ],
    },
    "GO100": {
        "host": "contabo14",
        "paths": [
            {"base": "/root/kis-autotrade-v4/report", "label": "리포트"},
            {"base": "/root/kis-autotrade-v4/reports", "label": "리포트"},
            {"base": "/root/kis-autotrade-v4/artifacts/go100", "label": "GO100 산출물",
             "include": ["latest.md", "report", "summary", "audit", "plan", ".html"]},
            {"base": "/root/kis-autotrade-v4/docs/go100", "label": "문서"},
            {"base": "/root/kis-autotrade-v4/docs/technical", "label": "기술문서"},
            {"base": "/root/kis-autotrade-v4/docs/reports", "label": "문서 리포트"},
            {"base": "/root/kis-autotrade-v4/docs/plans", "label": "기획문서"},
            {"base": "/root/kis-autotrade-v4/docs/plan", "label": "기획문서"},
            {"base": "/root/kis-autotrade-v4/docs/api", "label": "API 문서"},
            {"base": "/root/kis-autotrade-v4/docs/handover", "label": "인수인계"},
            {"base": "/root/kis-autotrade-v4/docs/operations", "label": "운영문서"},
            {"base": "/root/kis-autotrade-v4/docs/architecture", "label": "아키텍처"},
            {"base": "/root/kis-autotrade-v4/docs/design", "label": "설계문서"},
            {"base": "/root/kis-autotrade-v4/docs/agenda", "label": "아젠다"},
            {"base": "/root/kis-autotrade-v4/docs/features", "label": "기능명세"},
            {"base": "/root/kis-autotrade-v4/docs/analysis", "label": "분석문서"},
            {"base": "/root/kis-autotrade-v4/docs/whitepapers", "label": "백서"},
            {"base": "/root/kis-autotrade-v4/docs", "label": "문서",
             "include": ["GO100", "go100"],
             "exclude": [
                 "go100/", "technical/", "reports/", "plans/", "plan/", "api/",
                 "handover/", "operations/", "architecture/", "design/",
                 "agenda/", "features/", "analysis/", "whitepapers/",
                 "kis-api-portal/",
             ]},
        ],
    },
    "SF": {
        "host": "cafe24_114",
        "paths": [
            {"base": "/data/shortflow/docs", "label": "서비스 문서"},
        ],
    },
    "NTV2": {
        "host": "cafe24_114",
        "paths": [
            {"base": "/srv/newtalk-v2/docs", "label": "서비스 문서"},
        ],
    },
}

EXTENSIONS = {
    # 문서/리포트
    ".md", ".txt", ".html", ".htm", ".rst", ".pdf",
    # 데이터
    ".json", ".yaml", ".yml", ".toml", ".xml", ".csv",
    # 코드
    ".py", ".sh", ".sql", ".js", ".ts", ".tsx", ".jsx", ".css",
    # 설정/로그
    ".ini", ".cfg", ".conf", ".log",
    # 이미지 (브라우저 표시 가능)
    ".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".bmp", ".ico",
    # 오피스 문서 (다운로드 안내)
    ".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp",
    # 일반 첨부/아카이브/미디어 (직접 링크 열람 및 다운로드 폴백)
    ".xls", ".xlsm", ".doc", ".ppt",
    ".zip", ".tar", ".gz", ".tgz", ".7z", ".rar",
    ".mp3", ".wav", ".m4a", ".mp4", ".mov", ".webm",
}

# 바이너리 처리 대상 (텍스트로 읽지 않고 base64/raw)
BINARY_EXTENSIONS = {
    ".pdf",
    ".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".bmp", ".ico",
    ".docx", ".xlsx", ".xls", ".xlsm", ".pptx", ".odt", ".ods", ".odp",
    ".doc", ".ppt",
    ".zip", ".tar", ".gz", ".tgz", ".7z", ".rar",
    ".mp3", ".wav", ".m4a", ".mp4", ".mov", ".webm",
}

BASE64_PREVIEW_EXTENSIONS = {
    ".pdf",
    ".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".bmp", ".ico",
}

EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
DOCX_EXTENSIONS = {".docx"}
POWERPOINT_EXTENSIONS = {".pptx", ".odp"}
XML_OFFICE_EXTENSIONS = {".docx", ".pptx", ".odt", ".ods", ".odp"}
LEGACY_OFFICE_EXTENSIONS = {".doc", ".xls", ".ppt"}
SENSITIVE_PATH_MARKERS = (".env", "secrets", "credentials", "id_rsa")
SENSITIVE_EXTENSIONS = {".key", ".pem"}
AADS_APP_ROOT = "/app"
AADS_APP_REL_PREFIXES = ("docs/", "reports/", "app/static/", "scripts/", "tests/")

LOCAL_BASE_ALIASES = {
    "/app": ["/app", "/root/aads/aads-server"],
    "/app/docs": ["/app/docs", "/root/aads/aads-server/docs"],
    "/app/reports": ["/app/reports", "/root/aads/aads-server/reports"],
    "/app/app/static/docs": ["/app/app/static/docs", "/root/aads/aads-server/app/static/docs"],
    "/app/app/static/reports": ["/app/app/static/reports", "/root/aads/aads-server/app/static/reports"],
    "/app/app/static/preview": ["/app/app/static/preview", "/root/aads/aads-server/app/static/preview"],
    "/app/app/static/gallery": ["/app/app/static/gallery", "/root/aads/aads-server/app/static/gallery"],
}

PROJECT_FILE_HINTS = [
    ("GO100", re.compile(r"^(GO100[-_]|GO100\b|#?\d+.*GO100|.*상한가|.*백억)", re.IGNORECASE)),
    ("KIS", re.compile(r"^(KIS[-_]|KIS\b|.*자동매매)", re.IGNORECASE)),
    ("SF", re.compile(r"^(SF[-_]|ShortFlow\b|.*shortflow|.*숏폼)", re.IGNORECASE)),
    ("NTV2", re.compile(r"^(NTV2[-_]|NT[-_]|NewTalk\b|.*newtalk)", re.IGNORECASE)),
]

LEGACY_AADS_PROJECT_BASES = {
    "GO100": {
        "/app/docs": "/root/kis-autotrade-v4/docs",
        "/app/reports": "/root/kis-autotrade-v4/reports",
    },
    "KIS": {
        "/app/docs": "/root/kis-autotrade-v4/docs",
        "/app/reports": "/root/kis-autotrade-v4/docs",
    },
    "SF": {
        "/app/docs": "/data/shortflow/docs",
        "/app/reports": "/data/shortflow/docs",
    },
    "NTV2": {
        "/app/docs": "/srv/newtalk-v2/docs",
        "/app/reports": "/srv/newtalk-v2/docs",
    },
}


def _configured_base_paths(project: str) -> set[str]:
    config = SERVER_CONFIG.get(project) or {}
    paths = {str(Path(path_cfg["base"])) for path_cfg in config.get("paths", [])}
    if project == "AADS":
        paths.add(AADS_APP_ROOT)
    return paths


def _is_safe_relative_path(file_path: str) -> bool:
    normalized = file_path.replace("\\", "/")
    path = Path(normalized)
    if not normalized or path.is_absolute() or ".." in path.parts:
        return False
    parts = [part.lower() for part in normalized.split("/") if part]
    if any(part in SENSITIVE_PATH_MARKERS or part.startswith(".env") or part.startswith("id_rsa") for part in parts):
        return False
    if any(marker in normalized.lower() for marker in ("secrets", "credentials")):
        return False
    return Path(normalized).suffix.lower() not in SENSITIVE_EXTENSIONS


def _candidate_local_bases(base_path: str) -> list[Path]:
    normalized = str(Path(base_path))
    return [Path(p) for p in LOCAL_BASE_ALIASES.get(normalized, [normalized])]


def _basename(path: str) -> str:
    return path.replace("\\", "/").rsplit("/", 1)[-1]


def _project_hint_from_file_path(file_path: str) -> Optional[str]:
    name = _basename(file_path)
    for project, pattern in PROJECT_FILE_HINTS:
        if pattern.search(name):
            return project
    return None


def _candidate_file_paths_for_base(base_path: str, file_path: str) -> list[str]:
    normalized = file_path.replace("\\", "/").lstrip("/")
    candidates = [normalized]
    if normalized.startswith("docs/reports/"):
        candidates.append(normalized.removeprefix("docs/"))
        candidates.append(normalized.removeprefix("docs/reports/"))
    if normalized.startswith("reports/"):
        candidates.append(normalized.removeprefix("reports/"))
    elif base_path.endswith("/docs") or base_path.endswith("/app/docs"):
        candidates.append(f"reports/{normalized}")
    return list(dict.fromkeys([item for item in candidates if item]))


def _content_location_candidates(project: str, base_path: str, file_path: str) -> list[tuple[str, str, str]]:
    normalized_project = project.strip()
    normalized_base = str(Path(base_path))
    normalized_file = file_path.replace("\\", "/").lstrip("/")
    candidates: list[tuple[str, str, str]] = [(normalized_project, normalized_base, normalized_file)]

    hinted_project = _project_hint_from_file_path(normalized_file)
    if normalized_project == "AADS" and hinted_project and normalized_base in {"/app/docs", "/app/reports"}:
        hinted_base = LEGACY_AADS_PROJECT_BASES.get(hinted_project, {}).get(normalized_base)
        if hinted_base:
            candidates.append((hinted_project, hinted_base, normalized_file))

    projects = [normalized_project]
    if hinted_project and hinted_project not in projects:
        projects.append(hinted_project)

    for candidate_project in projects:
        for candidate_base in sorted(_configured_base_paths(candidate_project), key=len, reverse=True):
            for candidate_file in _candidate_file_paths_for_base(candidate_base, normalized_file):
                candidates.append((candidate_project, candidate_base, candidate_file))

    deduped: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for item in candidates:
        if item in seen:
            continue
        seen.add(item)
        deduped.append(item)
    return deduped


def _matches_path_filters(rel_path: str, *, include: list[str] | None, exclude: list[str] | None) -> bool:
    """Return whether a scanned relative path should be shown in 문서현황."""
    normalized = rel_path.replace("\\", "/")
    if exclude and any(ex in normalized for ex in exclude):
        return False
    if include and not any(inc in normalized for inc in include):
        return False
    return True


def _resolve_local_file(project: str, base_path: str, file_path: str) -> Path:
    normalized_base = str(Path(base_path))
    allowed = _configured_base_paths(project)
    if normalized_base not in allowed:
        raise HTTPException(400, "Unsupported base_path")
    if not _is_safe_relative_path(file_path):
        raise HTTPException(400, "Invalid file path")
    normalized_file = file_path.replace("\\", "/")
    if project == "AADS" and normalized_base == AADS_APP_ROOT and not normalized_file.startswith(AADS_APP_REL_PREFIXES):
        raise HTTPException(403, "File path is not allowed under /app")

    rel = Path(normalized_file)
    first_candidate: Path | None = None
    for base in _candidate_local_bases(normalized_base):
        base_resolved = base.resolve()
        candidate = (base_resolved / rel).resolve()
        try:
            candidate.relative_to(base_resolved)
        except ValueError:
            raise HTTPException(400, "Invalid file path")
        if first_candidate is None:
            first_candidate = candidate
        if candidate.exists() and candidate.is_file():
            return candidate

    return first_candidate or (Path(normalized_base) / rel)


async def _remote_file_exists(host: str, full_path: str) -> bool:
    quoted_path = shlex.quote(full_path)
    output = await _run_cmd(
        ["ssh", "-o", "ConnectTimeout=5", host, f"test -f {quoted_path} && printf exists"],
        timeout=8,
    )
    return output.strip() == "exists"


async def _resolve_content_location(project: str, base_path: str, file_path: str) -> tuple[str, str, str, str]:
    for candidate_project, candidate_base, candidate_file in _content_location_candidates(project, base_path, file_path):
        if not _is_safe_relative_path(candidate_file):
            continue
        candidate_config = SERVER_CONFIG.get(candidate_project)
        if not candidate_config:
            continue
        normalized_base = str(Path(candidate_base))
        if normalized_base not in _configured_base_paths(candidate_project):
            continue

        host = candidate_config["host"]
        full_path = f"{normalized_base.rstrip('/')}/{candidate_file}"
        if host is None:
            try:
                local_path = _resolve_local_file(candidate_project, normalized_base, candidate_file)
            except HTTPException:
                continue
            if local_path.exists() and local_path.is_file():
                return candidate_project, normalized_base, candidate_file, str(local_path)
        elif await _remote_file_exists(host, full_path):
            return candidate_project, normalized_base, candidate_file, full_path

    normalized_base = str(Path(base_path))
    normalized_file = file_path.replace("\\", "/").lstrip("/")
    return project, normalized_base, normalized_file, f"{normalized_base.rstrip('/')}/{normalized_file}"


def _excel_bytes_to_csv_text(raw: bytes, filename: str) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(f"openpyxl unavailable: {e}") from e

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            buf = io.StringIO()
            writer = csv.writer(buf)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else str(cell) for cell in row])
            parts.append(f"## Sheet: {sheet_name}\n{buf.getvalue().rstrip()}")
    finally:
        wb.close()
    return f"# {filename} CSV preview\n\n" + "\n\n".join(parts)


def _docx_bytes_to_text(raw: bytes, filename: str) -> str:
    try:
        import docx
    except Exception as e:
        logger.warning("project_doc_python_docx_unavailable", filename=filename, error=str(e))
        return _zip_office_bytes_to_text(raw, filename, ".docx")

    doc = docx.Document(io.BytesIO(raw))
    paragraphs = [p.text for p in doc.paragraphs if p.text]
    return f"# {filename} text preview\n\n" + "\n".join(paragraphs)


def _xml_text_nodes(raw_xml: bytes) -> list[str]:
    text = raw_xml.decode("utf-8", errors="replace")
    nodes = re.findall(r"<[^>/!:]+:?t(?:\s[^>]*)?>(.*?)</[^>:]+:?t>", text, flags=re.DOTALL)
    if not nodes:
        nodes = re.findall(r"<text:[^>]+>(.*?)</text:[^>]+>", text, flags=re.DOTALL)
    if not nodes:
        stripped = re.sub(r"<[^>]+>", " ", text)
        nodes = [stripped]
    values: list[str] = []
    for node in nodes:
        value = re.sub(r"<[^>]+>", " ", node)
        value = html.unescape(value)
        value = re.sub(r"\s+", " ", value).strip()
        if value:
            values.append(value)
    return values


def _zip_office_bytes_to_text(raw: bytes, filename: str, ext: str) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            names = zf.namelist()
            parts: list[str] = []
            if ext == ".pptx":
                slide_names = sorted(
                    [name for name in names if re.match(r"ppt/slides/slide\d+\.xml$", name)],
                    key=lambda value: int(re.search(r"slide(\d+)\.xml$", value).group(1)),  # type: ignore[union-attr]
                )
                for index, name in enumerate(slide_names, start=1):
                    slide_text = _xml_text_nodes(zf.read(name))
                    if slide_text:
                        parts.append(f"## Slide {index}\n" + "\n".join(slide_text))
            elif ext == ".docx":
                doc_names = [
                    name for name in names
                    if name == "word/document.xml"
                    or re.match(r"word/(header|footer)\d+\.xml$", name)
                ]
                for name in doc_names:
                    parts.extend(_xml_text_nodes(zf.read(name)))
            elif ext in {".odt", ".ods", ".odp"} and "content.xml" in names:
                parts.extend(_xml_text_nodes(zf.read("content.xml")))
            if not parts:
                raise ValueError("No previewable Office XML text found")
            label = {
                ".docx": "Word",
                ".pptx": "PowerPoint",
                ".odt": "OpenDocument Text",
                ".ods": "OpenDocument Sheet",
                ".odp": "OpenDocument Presentation",
            }.get(ext, "Office")
            preview = "\n\n".join(parts)
            return f"# {filename} {label} preview\n\n{preview[:200_000]}"
    except Exception as e:
        raise RuntimeError(f"Office XML preview failed: {e}") from e


def _legacy_office_bytes_to_text(raw: bytes, filename: str) -> str:
    chunks: list[str] = []
    seen: set[str] = set()

    for encoding in ("utf-16le", "utf-8", "cp949", "latin-1"):
        decoded = raw.decode(encoding, errors="ignore")
        for match in re.findall(r"[^\x00-\x08\x0b\x0c\x0e-\x1f\x7f]{4,}", decoded):
            value = re.sub(r"\s+", " ", match).strip()
            if len(value) < 4 or value in seen:
                continue
            if sum(ch.isalnum() or "\uac00" <= ch <= "\ud7a3" for ch in value) < 3:
                continue
            seen.add(value)
            chunks.append(value[:500])
            if len(chunks) >= 300:
                break
        if len(chunks) >= 300:
            break

    if not chunks:
        raise RuntimeError("No readable legacy Office text found")
    return (
        f"# {filename} legacy Office text preview\n\n"
        "구형 Office 바이너리에서 추출한 텍스트 미리보기입니다. 원본 서식은 다운로드 파일에서 확인하십시오.\n\n"
        + "\n".join(f"- {chunk}" for chunk in chunks)
    )


def _office_preview_format(ext: str) -> str:
    if ext in {".xlsx", ".xlsm", ".ods", ".xls"}:
        return "excel-csv" if ext in {".xlsx", ".xlsm"} else "office-text"
    if ext in {".docx", ".odt", ".doc"}:
        return "word-text"
    if ext in {".pptx", ".odp", ".ppt"}:
        return "powerpoint-text"
    return "office-text"


def _office_preview_response(
    *,
    project: str,
    file_path: str,
    full_path: str,
    content: str,
    source_mime_type: str,
    ext: str,
) -> dict:
    return {
        "project": project,
        "file_path": file_path,
        "full_path": full_path,
        "content": content,
        "size": len(content),
        "encoding": "text",
        "mime_type": "text/plain",
        "is_binary": False,
        "source_mime_type": source_mime_type,
        "converted_from": ext.lstrip("."),
        "format": _office_preview_format(ext),
    }


def _load_persistent_cache() -> Optional[dict]:
    """프로세스 재시작 후에도 이전 문서 목록을 즉시 재사용한다."""
    if _cache["data"]:
        return _cache["data"]
    try:
        if not PERSISTENT_CACHE_FILE.exists():
            return None
        data = json.loads(PERSISTENT_CACHE_FILE.read_text(encoding="utf-8"))
        if not isinstance(data, dict) or data.get("status") != "ok":
            return None
        _cache["data"] = data
        _cache["ts"] = int(data.get("scanned_at") or 0)
        return data
    except Exception as e:
        logger.warning("project_docs_cache_load_failed", path=str(PERSISTENT_CACHE_FILE), error=str(e))
        return None


def _save_persistent_cache(data: dict) -> None:
    """스캔 결과를 파일 캐시에 저장한다. 실패해도 API 응답은 유지한다."""
    try:
        PERSISTENT_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        PERSISTENT_CACHE_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        logger.warning("project_docs_cache_save_failed", path=str(PERSISTENT_CACHE_FILE), error=str(e))


def _file_key(doc: dict) -> tuple[str, str]:
    return (doc.get("base_path", ""), doc.get("path", ""))


def _file_signature(doc: dict) -> tuple[int, int]:
    return (int(doc.get("size") or 0), int(doc.get("modified") or 0))


def _previous_project(previous: Optional[dict], project: str) -> Optional[dict]:
    if not previous:
        return None
    for item in previous.get("projects") or []:
        if item.get("project") == project:
            return item
    return None


def _attach_delta(current: dict, previous: Optional[dict]) -> dict:
    """기존 목록과 비교해 이번 스캔에서 실제 변경된 파일 수를 표시한다."""
    prev_files = previous.get("files", []) if previous else []
    prev_map = {_file_key(doc): doc for doc in prev_files}
    curr_map = {_file_key(doc): doc for doc in current.get("files", [])}

    new_count = 0
    updated_count = 0
    unchanged_count = 0
    for key, doc in curr_map.items():
        prev_doc = prev_map.get(key)
        if not prev_doc:
            new_count += 1
        elif _file_signature(prev_doc) != _file_signature(doc):
            updated_count += 1
        else:
            unchanged_count += 1

    current["delta"] = {
        "new": new_count,
        "updated": updated_count,
        "removed": max(0, len(prev_map) - len(curr_map.keys() & prev_map.keys())),
        "unchanged": unchanged_count,
    }
    return current


async def _run_cmd(cmd: list[str], timeout: float = 10) -> str:
    """subprocess 실행 헬퍼."""
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        return stdout.decode("utf-8", errors="replace") if stdout else ""
    except (asyncio.TimeoutError, Exception) as e:
        logger.warning("cmd_failed", cmd=cmd[:3], error=str(e))
        return ""


async def _scan_local(base: str, exclude: list[str] | None = None, include: list[str] | None = None) -> list[dict]:
    """로컬 파일시스템 스캔."""
    results = []
    base_path = Path(base)
    if not base_path.exists():
        return results
    for p in sorted(base_path.rglob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() not in EXTENSIONS:
            continue
        rel = str(p.relative_to(base_path))
        if not _matches_path_filters(rel, include=include, exclude=exclude):
            continue
        stat = p.stat()
        results.append({
            "name": p.name,
            "path": rel,
            "size": stat.st_size,
            "modified": int(stat.st_mtime),
            "type": _classify(p.name, rel),
            "format": _detect_format(p.name),
        })
    return results


async def _scan_remote(host: str, base: str, exclude: list[str] | None = None, include: list[str] | None = None) -> list[dict]:
    """SSH로 원격 서버 스캔."""
    ext_pattern = " -o ".join(f'-name "*.{ext.lstrip(".")}"' for ext in EXTENSIONS)
    find_cmd = f'find {base} -type f \\( {ext_pattern} \\) -printf "%P\\t%s\\t%T@\\n" 2>/dev/null'
    output = await _run_cmd(["ssh", "-o", "ConnectTimeout=5", host, find_cmd], timeout=15)
    if not output.strip():
        return []
    results = []
    for line in output.strip().split("\n"):
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        rel_path, size_str, mtime_str = parts[0], parts[1], parts[2]
        name = rel_path.rsplit("/", 1)[-1] if "/" in rel_path else rel_path
        if not _matches_path_filters(rel_path, include=include, exclude=exclude):
            continue
        results.append({
            "name": name,
            "path": rel_path,
            "size": int(size_str) if size_str.isdigit() else 0,
            "modified": int(float(mtime_str)) if mtime_str else 0,
            "type": _classify(name, rel_path),
            "format": _detect_format(name),
        })
    return sorted(results, key=lambda x: x["path"])


def _classify(name: str, path: str) -> str:
    """문서 유형 분류."""
    nl = name.lower()
    pl = path.lower()

    if any(key in pl for key in ("contract", "contracts", "agreement", "agreements", "계약")) or \
            any(key in nl for key in ("contract", "agreement", "계약", "근로계약", "입점계약", "프리랜서")):
        return "contract"
    if any(key in pl for key in ("ceo-documents", "directive", "directives", "policy", "rule")) or \
            any(key in nl for key in ("directive", "directives", "policy", "rules")):
        return "directive"
    if "handover" in nl or "handover" in pl:
        return "handover"
    if any(key in pl for key in ("changelog", "release-note", "release_note", "history")) or \
            any(key in nl for key in ("changelog", "release-note", "release_note", "history")):
        return "changelog"
    if any(key in pl for key in ("report", "result", "retrospective", "postmortem")) or \
            any(key in nl for key in ("report", "result", "retrospective", "postmortem")):
        return "report"
    if any(key in pl for key in ("qa", "test", "verification", "benchmark")) or \
            any(key in nl for key in ("qa", "test", "verification", "benchmark")):
        return "qa"
    if any(key in pl for key in ("api", "openapi", "swagger")) or \
            any(key in nl for key in ("api", "openapi", "swagger")):
        return "api"
    if any(key in pl for key in ("architecture", "system-design", "design", "technical", "tech")) or \
            any(key in nl for key in ("architecture", "design", "technical", "tech")):
        return "architecture"
    if any(key in pl for key in ("runbook", "deploy", "deployment", "operation", "ops", "playbook", "troubleshoot")) or \
            any(key in nl for key in ("runbook", "deploy", "deployment", "operation", "ops", "playbook", "troubleshoot")):
        return "runbook"
    if any(key in pl for key in ("plan", "roadmap", "proposal", "spec", "prd", "layout")) or \
            any(key in nl for key in ("plan", "roadmap", "proposal", "spec", "prd")):
        return "plan"
    if any(key in pl for key in ("status", "incident", "issue", "summary")) or \
            any(key in nl for key in ("status", "incident", "issue", "summary")):
        return "status"
    if any(key in pl for key in ("lesson", "knowledge", "guide", "manual", "faq", "tutorial")) or \
            any(key in nl for key in ("lesson", "knowledge", "guide", "manual", "faq", "tutorial")):
        return "knowledge"
    if nl.endswith(".sql") or any(key in pl for key in ("schema", "migration", "erd", "ddl")) or \
            any(key in nl for key in ("schema", "migration", "erd", "ddl")):
        return "schema"
    if nl.endswith((".py", ".sh")):
        return "script"
    if nl.endswith((".json", ".yaml", ".yml")) or \
            any(key in pl for key in ("config", "settings", "compose", "env")) or \
            any(key in nl for key in ("config", "settings", "compose", "env")):
        return "config"
    return "doc"


def _detect_format(name: str) -> str:
    """파일 확장자 기반 포맷 분류."""
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    _FORMAT_MAP = {
        # 문서
        "md": "markdown", "txt": "text", "html": "html", "htm": "html", "rst": "rst", "pdf": "pdf",
        # 데이터
        "json": "json", "yaml": "yaml", "yml": "yaml", "toml": "toml", "xml": "xml", "csv": "csv",
        # 코드
        "py": "python", "sh": "shell", "sql": "sql",
        "js": "javascript", "ts": "typescript", "tsx": "typescript", "jsx": "javascript",
        "css": "css", "ini": "config", "cfg": "config", "conf": "config", "log": "log",
        # 이미지
        "png": "image", "jpg": "image", "jpeg": "image", "gif": "image",
        "svg": "image", "webp": "image", "bmp": "image", "ico": "image",
        # 오피스
        "docx": "word", "doc": "word",
        "xlsx": "excel", "xls": "excel", "xlsm": "excel",
        "pptx": "powerpoint", "ppt": "powerpoint",
        "odt": "word", "ods": "excel", "odp": "powerpoint",
    }
    return _FORMAT_MAP.get(ext, "other")


async def _scan_project(project: str, config: dict, previous: Optional[dict] = None) -> dict:
    """프로젝트 1개 스캔."""
    host = config["host"]
    all_docs = []
    for path_cfg in config["paths"]:
        base = path_cfg["base"]
        exclude = path_cfg.get("exclude")
        include = path_cfg.get("include")
        label = path_cfg["label"]
        if host is None:
            docs = await _scan_local(base, exclude, include)
        else:
            docs = await _scan_remote(host, base, exclude, include)
        for d in docs:
            d["base_path"] = base
            d["label"] = label
            d["full_path"] = f"{base.rstrip('/')}/{d['path']}"
        all_docs.extend(docs)

    # 같은 실파일이 여러 base_path에서 중복 노출되지 않도록 정규화 dedupe
    deduped = []
    seen = set()
    for d in all_docs:
        key = (d.get("base_path", ""), d.get("path", ""))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(d)
    return _attach_delta({
        "project": project,
        "host": host or "localhost",
        "total": len(deduped),
        "files": deduped,
    }, previous)


@router.get("/project-docs/scan")
async def scan_all_docs(force: bool = Query(False, description="캐시 무시하고 재스캔")):
    """전 서버 문서 스캔.

    - 일반 호출: 5분 메모리 캐시 우선, 프로세스 재시작 후에는 파일 캐시 우선.
    - 강제 호출: 기존 캐시와 비교해 new/updated/removed만 delta로 표시한다.
    """
    now = time.time()
    previous = _load_persistent_cache()
    if not force and previous and (now - _cache["ts"]) < CACHE_TTL:
        resp = dict(previous)
        resp["cache_hit"] = True
        resp["cache_age_sec"] = int(now - _cache["ts"])
        resp["cache_mode"] = "memory" if _cache["data"] is previous else "file"
        return resp

    tasks = [_scan_project(proj, cfg, _previous_project(previous, proj)) for proj, cfg in SERVER_CONFIG.items()]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    projects = []
    total = 0
    for r in results:
        if isinstance(r, Exception):
            logger.error("scan_error", error=str(r))
            continue
        projects.append(r)
        total += r["total"]

    resp = {
        "status": "ok",
        "total": total,
        "projects": projects,
        "scanned_at": int(now),
        "cache_hit": False,
        "cache_mode": "incremental" if previous else "full",
        "delta": {
            "new": sum((p.get("delta") or {}).get("new", 0) for p in projects),
            "updated": sum((p.get("delta") or {}).get("updated", 0) for p in projects),
            "removed": sum((p.get("delta") or {}).get("removed", 0) for p in projects),
            "unchanged": sum((p.get("delta") or {}).get("unchanged", 0) for p in projects),
        },
    }
    _cache["data"] = resp
    _cache["ts"] = now
    _save_persistent_cache(resp)
    return resp


@router.get("/project-docs/content")
async def get_doc_content(
    project: str = Query(..., description="프로젝트명 (AADS/KIS/GO100/SF/NTV2)"),
    base_path: str = Query(..., description="base_path (스캔 결과에서 제공)"),
    file_path: str = Query(..., description="파일 상대 경로"),
):
    """문서 내용 조회."""
    config = SERVER_CONFIG.get(project)
    if not config:
        raise HTTPException(400, f"Unknown project: {project}")

    # 경로 검증 (traversal 방지)
    if not _is_safe_relative_path(file_path):
        raise HTTPException(400, "Invalid file path")

    project, base_path, file_path, full_path = await _resolve_content_location(project, base_path, file_path)
    config = SERVER_CONFIG.get(project)
    if not config:
        raise HTTPException(400, f"Unknown project: {project}")
    host = config["host"]

    # 확장자 기반 바이너리 판별
    ext = ("." + file_path.rsplit(".", 1)[-1].lower()) if "." in file_path else ""
    is_binary = ext in BINARY_EXTENSIONS
    mime_type = mimetypes.guess_type(file_path)[0] or "application/octet-stream"

    if host is None:
        # 로컬 파일
        p = _resolve_local_file(project, base_path, file_path)
        full_path = str(p)
        if not p.exists() or not p.is_file():
            raise HTTPException(404, "File not found")
        size_bytes = p.stat().st_size
        if ext in EXCEL_EXTENSIONS:
            raw = p.read_bytes()
            try:
                content = _excel_bytes_to_csv_text(raw, p.name)
                return {
                    "project": project,
                    "file_path": file_path,
                    "full_path": full_path,
                    "content": content,
                    "size": len(content),
                    "encoding": "text",
                    "mime_type": "text/csv",
                    "is_binary": False,
                    "source_mime_type": mime_type,
                    "converted_from": ext.lstrip("."),
                    "format": "excel-csv",
                }
            except Exception as e:
                logger.warning("project_doc_excel_preview_failed", path=full_path, error=str(e))
                is_binary = True
        if ext in DOCX_EXTENSIONS:
            raw = p.read_bytes()
            try:
                content = _docx_bytes_to_text(raw, p.name)
                return _office_preview_response(
                    project=project,
                    file_path=file_path,
                    full_path=full_path,
                    content=content,
                    source_mime_type=mime_type,
                    ext=ext,
                )
            except Exception as e:
                logger.warning("project_doc_docx_preview_failed", path=full_path, error=str(e))
                is_binary = True
        if ext in (XML_OFFICE_EXTENSIONS - EXCEL_EXTENSIONS - DOCX_EXTENSIONS):
            raw = p.read_bytes()
            try:
                content = _zip_office_bytes_to_text(raw, p.name, ext)
                return _office_preview_response(
                    project=project,
                    file_path=file_path,
                    full_path=full_path,
                    content=content,
                    source_mime_type=mime_type,
                    ext=ext,
                )
            except Exception as e:
                logger.warning("project_doc_xml_office_preview_failed", path=full_path, error=str(e))
                is_binary = True
        if ext in LEGACY_OFFICE_EXTENSIONS:
            raw = p.read_bytes()
            try:
                content = _legacy_office_bytes_to_text(raw, p.name)
                return _office_preview_response(
                    project=project,
                    file_path=file_path,
                    full_path=full_path,
                    content=content,
                    source_mime_type=mime_type,
                    ext=ext,
                )
            except Exception as e:
                logger.warning("project_doc_legacy_office_preview_failed", path=full_path, error=str(e))
                is_binary = True
        # 텍스트 1MB / 바이너리 10MB 한도
        max_size = 10_000_000 if is_binary else 1_000_000
        if size_bytes > max_size:
            raise HTTPException(413, f"File too large (>{max_size // 1_000_000}MB)")
        if is_binary:
            raw = p.read_bytes()
            content = base64.b64encode(raw).decode("ascii")
        else:
            content = p.read_text(encoding="utf-8", errors="replace")
    else:
        # 원격 파일
        normalized_base = str(Path(base_path))
        if normalized_base not in _configured_base_paths(project):
            raise HTTPException(400, "Unsupported base_path")
        if ext in EXCEL_EXTENSIONS:
            quoted_full_path = shlex.quote(full_path)
            b64_output = await _run_cmd(
                ["ssh", "-o", "ConnectTimeout=5", host, f"base64 -w0 {quoted_full_path} 2>/dev/null"],
                timeout=15,
            )
            if not b64_output:
                raise HTTPException(404, "File not found or empty")
            raw = base64.b64decode(b64_output.strip())
            try:
                content = _excel_bytes_to_csv_text(raw, Path(file_path).name)
                return {
                    "project": project,
                    "file_path": file_path,
                    "full_path": full_path,
                    "content": content,
                    "size": len(content),
                    "encoding": "text",
                    "mime_type": "text/csv",
                    "is_binary": False,
                    "source_mime_type": mime_type,
                    "converted_from": ext.lstrip("."),
                    "format": "excel-csv",
                }
            except Exception as e:
                logger.warning("project_doc_remote_excel_preview_failed", path=full_path, error=str(e))
                is_binary = True
        if ext in DOCX_EXTENSIONS:
            quoted_full_path = shlex.quote(full_path)
            b64_output = await _run_cmd(
                ["ssh", "-o", "ConnectTimeout=5", host, f"base64 -w0 {quoted_full_path} 2>/dev/null"],
                timeout=15,
            )
            if not b64_output:
                raise HTTPException(404, "File not found or empty")
            raw = base64.b64decode(b64_output.strip())
            try:
                content = _docx_bytes_to_text(raw, Path(file_path).name)
                return _office_preview_response(
                    project=project,
                    file_path=file_path,
                    full_path=full_path,
                    content=content,
                    source_mime_type=mime_type,
                    ext=ext,
                )
            except Exception as e:
                logger.warning("project_doc_remote_docx_preview_failed", path=full_path, error=str(e))
                is_binary = True
        if ext in (XML_OFFICE_EXTENSIONS - EXCEL_EXTENSIONS - DOCX_EXTENSIONS):
            quoted_full_path = shlex.quote(full_path)
            b64_output = await _run_cmd(
                ["ssh", "-o", "ConnectTimeout=5", host, f"base64 -w0 {quoted_full_path} 2>/dev/null"],
                timeout=15,
            )
            if not b64_output:
                raise HTTPException(404, "File not found or empty")
            raw = base64.b64decode(b64_output.strip())
            try:
                content = _zip_office_bytes_to_text(raw, Path(file_path).name, ext)
                return _office_preview_response(
                    project=project,
                    file_path=file_path,
                    full_path=full_path,
                    content=content,
                    source_mime_type=mime_type,
                    ext=ext,
                )
            except Exception as e:
                logger.warning("project_doc_remote_xml_office_preview_failed", path=full_path, error=str(e))
                is_binary = True
        if ext in LEGACY_OFFICE_EXTENSIONS:
            quoted_full_path = shlex.quote(full_path)
            b64_output = await _run_cmd(
                ["ssh", "-o", "ConnectTimeout=5", host, f"base64 -w0 {quoted_full_path} 2>/dev/null"],
                timeout=15,
            )
            if not b64_output:
                raise HTTPException(404, "File not found or empty")
            raw = base64.b64decode(b64_output.strip())
            try:
                content = _legacy_office_bytes_to_text(raw, Path(file_path).name)
                return _office_preview_response(
                    project=project,
                    file_path=file_path,
                    full_path=full_path,
                    content=content,
                    source_mime_type=mime_type,
                    ext=ext,
                )
            except Exception as e:
                logger.warning("project_doc_remote_legacy_office_preview_failed", path=full_path, error=str(e))
                is_binary = True
        if is_binary:
            # SSH base64 인코딩으로 바이너리 안전 전송
            quoted_full_path = shlex.quote(full_path)
            b64_output = await _run_cmd(
                ["ssh", "-o", "ConnectTimeout=5", host, f"base64 -w0 {quoted_full_path} 2>/dev/null"],
                timeout=15,
            )
            if not b64_output:
                raise HTTPException(404, "File not found or empty")
            content = b64_output.strip()
        else:
            quoted_full_path = shlex.quote(full_path)
            content = await _run_cmd(
                ["ssh", "-o", "ConnectTimeout=5", host, f"cat {quoted_full_path}"],
                timeout=10,
            )
            if not content:
                raise HTTPException(404, "File not found or empty")

    return {
        "project": project,
        "file_path": file_path,
        "full_path": full_path,
        "content": content,
        "size": len(content),
        "encoding": "base64" if is_binary else "text",
        "mime_type": mime_type,
        "is_binary": is_binary,
        "format": "binary" if is_binary and ext not in BASE64_PREVIEW_EXTENSIONS else _detect_format(file_path),
    }
