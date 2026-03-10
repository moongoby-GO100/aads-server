"""
AADS-190: 데이터 내보내기 도구.
CEO 채팅에서 쿼리 결과를 Excel/CSV/PDF로 변환 → 다운로드 링크 제공.

지원 포맷:
- CSV: 기본, 경량
- Excel (xlsx): openpyxl 기반, 자동 열 너비 + 헤더 스타일
- PDF: weasyprint 기반 HTML→PDF 변환

파일은 /tmp/aads_exports/에 저장, nginx에서 /exports/ 경로로 서빙.
"""
from __future__ import annotations

import csv
import io
import logging
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)

_EXPORT_DIR = Path(os.getenv("AADS_EXPORT_DIR", "/var/www/aads_exports"))
_EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# 다운로드 URL 베이스 (nginx /exports/ 경로로 서빙)
_EXPORT_URL_BASE = os.getenv("AADS_EXPORT_URL", "https://aads.newtalk.kr/exports")


def _now_kst() -> str:
    return datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y%m%d_%H%M%S")


async def export_to_csv(
    rows: List[Dict[str, Any]],
    filename: Optional[str] = None,
    title: Optional[str] = None,
) -> Dict[str, Any]:
    """데이터를 CSV로 내보내기."""
    if not rows:
        return {"error": "내보낼 데이터가 없습니다"}

    fname = filename or f"export_{_now_kst()}_{uuid.uuid4().hex[:6]}.csv"
    if not fname.endswith(".csv"):
        fname += ".csv"

    filepath = _EXPORT_DIR / fname
    columns = list(rows[0].keys())

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

    size = filepath.stat().st_size
    logger.info(f"export_csv: {fname} ({size} bytes, {len(rows)} rows)")

    return {
        "format": "csv",
        "filename": fname,
        "path": str(filepath),
        "url": f"{_EXPORT_URL_BASE}/{fname}",
        "rows": len(rows),
        "columns": columns,
        "size_bytes": size,
    }


async def export_to_excel(
    rows: List[Dict[str, Any]],
    filename: Optional[str] = None,
    title: Optional[str] = None,
    sheet_name: str = "Data",
) -> Dict[str, Any]:
    """데이터를 Excel(xlsx)로 내보내기. 자동 열 너비 + 헤더 스타일."""
    if not rows:
        return {"error": "내보낼 데이터가 없습니다"}

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"error": "openpyxl이 설치되지 않았습니다. pip install openpyxl"}

    fname = filename or f"export_{_now_kst()}_{uuid.uuid4().hex[:6]}.xlsx"
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"

    filepath = _EXPORT_DIR / fname
    columns = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 제목 행 (선택)
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        start_row = 3

    # 헤더
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    # 자동 열 너비
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in rows[:100]:  # 처음 100행만 측정
            val = str(row.get(col_name, ""))
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"{chr(64 + (col_idx-1)//26)}{chr(64 + (col_idx-1)%26 + 1)}"].width = max_len + 3

    wb.save(filepath)
    size = filepath.stat().st_size
    logger.info(f"export_excel: {fname} ({size} bytes, {len(rows)} rows)")

    return {
        "format": "xlsx",
        "filename": fname,
        "path": str(filepath),
        "url": f"{_EXPORT_URL_BASE}/{fname}",
        "rows": len(rows),
        "columns": columns,
        "size_bytes": size,
    }


async def export_to_pdf(
    rows: List[Dict[str, Any]],
    filename: Optional[str] = None,
    title: Optional[str] = None,
) -> Dict[str, Any]:
    """데이터를 PDF로 내보내기 (HTML 테이블 → PDF)."""
    if not rows:
        return {"error": "내보낼 데이터가 없습니다"}

    fname = filename or f"export_{_now_kst()}_{uuid.uuid4().hex[:6]}.pdf"
    if not fname.endswith(".pdf"):
        fname += ".pdf"

    filepath = _EXPORT_DIR / fname
    columns = list(rows[0].keys())

    # HTML 생성
    title_html = f"<h1>{title}</h1>" if title else ""
    header_cells = "".join(f"<th>{c}</th>" for c in columns)
    body_rows = ""
    for row in rows:
        cells = "".join(f"<td>{row.get(c, '')}</td>" for c in columns)
        body_rows += f"<tr>{cells}</tr>"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: 'Noto Sans KR', sans-serif; margin: 20px; font-size: 10px; }}
  h1 {{ font-size: 16px; margin-bottom: 10px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #4472C4; color: white; padding: 6px 8px; text-align: left; }}
  td {{ border-bottom: 1px solid #ddd; padding: 4px 8px; }}
  tr:nth-child(even) {{ background: #f9f9f9; }}
  .footer {{ margin-top: 10px; font-size: 8px; color: #888; }}
</style></head>
<body>
  {title_html}
  <table><thead><tr>{header_cells}</tr></thead>
  <tbody>{body_rows}</tbody></table>
  <div class="footer">Generated by AADS · {_now_kst()} KST · {len(rows)} rows</div>
</body></html>"""

    # weasyprint로 PDF 변환 시도, 없으면 HTML로 폴백
    try:
        from weasyprint import HTML
        HTML(string=html).write_pdf(filepath)
    except ImportError:
        # weasyprint 없으면 HTML 파일로 저장
        fname = fname.replace(".pdf", ".html")
        filepath = _EXPORT_DIR / fname
        filepath.write_text(html, encoding="utf-8")
        logger.warning("export_pdf: weasyprint 미설치, HTML로 폴백")

    size = filepath.stat().st_size
    fmt = "pdf" if fname.endswith(".pdf") else "html"
    logger.info(f"export_{fmt}: {fname} ({size} bytes, {len(rows)} rows)")

    return {
        "format": fmt,
        "filename": fname,
        "path": str(filepath),
        "url": f"{_EXPORT_URL_BASE}/{fname}",
        "rows": len(rows),
        "columns": columns,
        "size_bytes": size,
    }


async def export_data(
    data: List[Dict[str, Any]],
    fmt: str = "xlsx",
    filename: Optional[str] = None,
    title: Optional[str] = None,
) -> Dict[str, Any]:
    """
    통합 내보내기 도구.

    Args:
        data: 내보낼 데이터 (dict 리스트)
        fmt: 포맷 — csv, xlsx, pdf
        filename: 파일명 (선택)
        title: 제목 (선택)
    """
    fmt = fmt.lower().strip()
    if fmt == "csv":
        return await export_to_csv(data, filename, title)
    elif fmt in ("xlsx", "excel"):
        return await export_to_excel(data, filename, title)
    elif fmt == "pdf":
        return await export_to_pdf(data, filename, title)
    else:
        return {"error": f"지원 포맷: csv, xlsx, pdf (입력: {fmt})"}
