#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
제조업 표준원가 관리 엑셀 양식 생성기
생성 파일: /tmp/standard_cost_template.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/tmp/standard_cost_template.xlsx"

# ── 색상 정의 ──────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL    = PatternFill("solid", fgColor="2E75B6")
FORMULA_FILL  = PatternFill("solid", fgColor="FFFACD")
TOTAL_FILL    = PatternFill("solid", fgColor="D6E4F0")
GUIDE_FILL    = PatternFill("solid", fgColor="EBF3FB")

WHITE_FONT   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="맑은 고딕", bold=True, size=10)
NORMAL_FONT  = Font(name="맑은 고딕", size=10)
GUIDE_FONT   = Font(name="맑은 고딕", italic=True, size=9, color="1F4E79")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

NUM_FMT   = '#,##0'
FLOAT_FMT = '#,##0.00'


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_cell(cell, text):
    cell.value = text
    cell.fill  = HEADER_FILL
    cell.font  = WHITE_FONT
    cell.alignment = CENTER
    cell.border = thin_border()

def style_title_row(ws, row, ncols, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.fill  = TITLE_FILL
    cell.font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
    cell.alignment = CENTER
    cell.border = medium_border()

def style_guide_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.fill  = GUIDE_FILL
    cell.font  = GUIDE_FONT
    cell.alignment = LEFT
    cell.border = thin_border()
    ws.row_dimensions[row].height = 20

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def dc(cell, value, fmt=None, formula=False, total=False, align=None):
    """데이터 셀 한 번에 스타일링"""
    cell.value = value
    if formula:
        cell.fill = FORMULA_FILL
    elif total:
        cell.fill = TOTAL_FILL
        cell.font = BOLD_FONT
    else:
        cell.font = NORMAL_FONT
    if fmt:
        cell.number_format = fmt
    cell.border = thin_border()
    if align:
        cell.alignment = align
    elif isinstance(value, (int, float)):
        cell.alignment = RIGHT
    else:
        cell.alignment = LEFT


# ══════════════════════════════════════════════════════════
# Sheet 1: BOM
# ══════════════════════════════════════════════════════════
def build_bom(wb):
    ws = wb.create_sheet("BOM")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    NCOLS = 11
    HEADERS = [
        "순번", "제품코드", "제품명", "자재코드", "자재명",
        "규격", "단위", "소요수량", "표준단가(원)", "재료비소계(원)", "비고"
    ]

    style_title_row(ws, 1, NCOLS, "BOM (자재명세서) — 구매부서 입력")
    ws.row_dimensions[1].height = 32

    style_guide_row(ws, 2, NCOLS,
        "  ※ 작성 안내: 자재코드, 자재명, 규격, 단위, 소요수량, 표준단가를 입력하세요. "
        "재료비소계(연노랑)는 소요수량×표준단가로 자동 계산됩니다.")

    ws.row_dimensions[3].height = 6

    ws.row_dimensions[4].height = 28
    for col, h in enumerate(HEADERS, 1):
        style_header_cell(ws.cell(row=4, column=col), h)

    samples = [
        (1, "P-001", "제품A", "M-001", "철판 SS400",     "2.0T×1000×2000", "장", 2,   45000),
        (2, "P-001", "제품A", "M-002", "볼트 M10×30",    "M10 L30",         "개", 8,     350),
        (3, "P-002", "제품B", "M-003", "알루미늄 6063",  "T5×50×3000",      "m",  5,  12000),
    ]

    for r, row in enumerate(samples, 5):
        seq, pcode, pname, mcode, mname, spec, unit, qty, price = row
        ws.row_dimensions[r].height = 18
        dc(ws.cell(r,1),  seq,   align=CENTER)
        dc(ws.cell(r,2),  pcode, align=CENTER)
        dc(ws.cell(r,3),  pname, align=LEFT)
        dc(ws.cell(r,4),  mcode, align=CENTER)
        dc(ws.cell(r,5),  mname, align=LEFT)
        dc(ws.cell(r,6),  spec,  align=LEFT)
        dc(ws.cell(r,7),  unit,  align=CENTER)
        dc(ws.cell(r,8),  qty,   fmt=FLOAT_FMT)
        dc(ws.cell(r,9),  price, fmt=NUM_FMT)
        # 수식: 재료비소계
        c = ws.cell(r, 10)
        c.value = f"=H{r}*I{r}"
        c.number_format = NUM_FMT
        c.fill = FORMULA_FILL
        c.alignment = RIGHT
        c.border = thin_border()
        c.font = NORMAL_FONT
        dc(ws.cell(r, 11), "", align=LEFT)

    # 합계행
    tr = 5 + len(samples)
    ws.row_dimensions[tr].height = 22
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=9)
    c = ws.cell(tr, 1)
    c.value="합  계"; c.fill=TOTAL_FILL; c.font=BOLD_FONT; c.alignment=CENTER; c.border=thin_border()
    for col in range(2, 10):
        ws.cell(tr, col).fill=TOTAL_FILL; ws.cell(tr,col).border=thin_border()
    c = ws.cell(tr, 10)
    c.value=f"=SUM(J5:J{tr-1})"; c.number_format=NUM_FMT
    c.fill=TOTAL_FILL; c.font=BOLD_FONT; c.alignment=RIGHT; c.border=thin_border()
    ws.cell(tr, 11).fill=TOTAL_FILL; ws.cell(tr,11).border=thin_border()

    set_col_widths(ws, [6, 12, 14, 12, 18, 18, 6, 10, 14, 14, 16])


# ══════════════════════════════════════════════════════════
# Sheet 2: 공정목록
# ══════════════════════════════════════════════════════════
def build_process_list(wb):
    ws = wb.create_sheet("공정목록")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    NCOLS = 6
    HEADERS = ["순번", "공정번호", "공정명", "담당부서", "공정설명", "비고"]

    style_title_row(ws, 1, NCOLS, "공정목록 — 생산관리팀 관리")
    ws.row_dimensions[1].height = 32

    style_guide_row(ws, 2, NCOLS,
        "  ※ 작성 안내: 공정번호, 공정명, 담당부서, 공정설명을 등록하세요. "
        "공정별원가입력 시트에서 참조됩니다.")

    ws.row_dimensions[3].height = 6

    ws.row_dimensions[4].height = 28
    for col, h in enumerate(HEADERS, 1):
        style_header_cell(ws.cell(row=4, column=col), h)

    samples = [
        (1, "PR-10", "원자재투입",  "자재창고",    "원자재 규격 확인 및 생산라인 투입"),
        (2, "PR-20", "가공",        "가공1팀",     "CNC 가공 및 프레스 성형"),
        (3, "PR-30", "조립",        "조립팀",      "부품 조립 및 체결"),
        (4, "PR-40", "검사",        "품질관리팀",  "치수 검사 및 외관 검사"),
        (5, "PR-50", "포장",        "포장팀",      "완제품 포장 및 라벨링"),
    ]

    for r, row in enumerate(samples, 5):
        seq, pno, pname, dept, desc = row
        ws.row_dimensions[r].height = 18
        dc(ws.cell(r,1), seq,   align=CENTER)
        dc(ws.cell(r,2), pno,   align=CENTER)
        dc(ws.cell(r,3), pname, align=LEFT)
        dc(ws.cell(r,4), dept,  align=CENTER)
        dc(ws.cell(r,5), desc,  align=LEFT)
        dc(ws.cell(r,6), "",    align=LEFT)

    set_col_widths(ws, [6, 12, 14, 14, 36, 16])


# ══════════════════════════════════════════════════════════
# Sheet 3: 공정별원가입력
# ══════════════════════════════════════════════════════════
def build_process_cost(wb):
    ws = wb.create_sheet("공정별원가입력")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    NCOLS = 12
    HEADERS = [
        "순번", "제품코드", "제품명", "공정번호", "공정명",
        "직접재료비(원)", "작업시간(H)", "시간당임률(원)",
        "직접노무비(원)", "제조간접비(원)", "공정원가합계(원)", "비고"
    ]

    style_title_row(ws, 1, NCOLS, "공정별 원가 입력 — 생산부서 입력")
    ws.row_dimensions[1].height = 32

    style_guide_row(ws, 2, NCOLS,
        "  ※ 작성 안내: 직접재료비, 작업시간, 시간당임률, 제조간접비를 입력하세요. "
        "직접노무비(=작업시간×시간당임률)·공정원가합계는 자동 계산됩니다.")

    ws.row_dimensions[3].height = 6

    ws.row_dimensions[4].height = 28
    for col, h in enumerate(HEADERS, 1):
        style_header_cell(ws.cell(row=4, column=col), h)

    # F=직접재료비, G=작업시간, H=시간당임률, I=직접노무비(수식), J=제조간접비, K=합계(수식)
    samples = [
        (1, "P-001", "제품A", "PR-10", "원자재투입",  90000, 0.5, 25000, 12500),
        (2, "P-001", "제품A", "PR-20", "가공",         15000, 2.0, 30000, 60000),
        (3, "P-001", "제품A", "PR-30", "조립",              0, 1.5, 28000, 42000),
        (4, "P-001", "제품A", "PR-40", "검사",              0, 0.5, 35000, 17500),
        (5, "P-002", "제품B", "PR-20", "가공",          20000, 3.0, 30000, 90000),
    ]

    for r, row in enumerate(samples, 5):
        seq, pcode, pname, prno, prname, mat, hrs, rate, overhead = row
        ws.row_dimensions[r].height = 18
        dc(ws.cell(r,1), seq,   align=CENTER)
        dc(ws.cell(r,2), pcode, align=CENTER)
        dc(ws.cell(r,3), pname, align=LEFT)
        dc(ws.cell(r,4), prno,  align=CENTER)
        dc(ws.cell(r,5), prname,align=LEFT)
        dc(ws.cell(r,6), mat,   fmt=NUM_FMT)
        dc(ws.cell(r,7), hrs,   fmt=FLOAT_FMT)
        dc(ws.cell(r,8), rate,  fmt=NUM_FMT)
        # I: 직접노무비 수식
        ci = ws.cell(r, 9)
        ci.value=f"=G{r}*H{r}"; ci.number_format=NUM_FMT
        ci.fill=FORMULA_FILL; ci.alignment=RIGHT; ci.border=thin_border(); ci.font=NORMAL_FONT
        dc(ws.cell(r,10), overhead, fmt=NUM_FMT)
        # K: 공정원가합계 수식
        ck = ws.cell(r, 11)
        ck.value=f"=F{r}+I{r}+J{r}"; ck.number_format=NUM_FMT
        ck.fill=FORMULA_FILL; ck.alignment=RIGHT; ck.border=thin_border(); ck.font=NORMAL_FONT
        dc(ws.cell(r,12), "", align=LEFT)

    # 합계행
    tr = 5 + len(samples)
    ws.row_dimensions[tr].height = 22
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
    c = ws.cell(tr, 1)
    c.value="합  계"; c.fill=TOTAL_FILL; c.font=BOLD_FONT; c.alignment=CENTER; c.border=thin_border()
    for col in range(2, 6):
        ws.cell(tr,col).fill=TOTAL_FILL; ws.cell(tr,col).border=thin_border()

    sum_map = {6:"F", 7:"G", 9:"I", 10:"J", 11:"K"}
    for col, letter in sum_map.items():
        c = ws.cell(tr, col)
        c.value = f"=SUM({letter}5:{letter}{tr-1})"
        c.number_format = FLOAT_FMT if col==7 else NUM_FMT
        c.fill=TOTAL_FILL; c.font=BOLD_FONT; c.alignment=RIGHT; c.border=thin_border()
    for col in [8, 12]:
        ws.cell(tr,col).fill=TOTAL_FILL; ws.cell(tr,col).border=thin_border()

    set_col_widths(ws, [6, 12, 14, 12, 14, 14, 11, 14, 14, 14, 16, 16])


# ══════════════════════════════════════════════════════════
# Sheet 4: 표준원가집계
# ══════════════════════════════════════════════════════════
def build_cost_summary(wb):
    ws = wb.create_sheet("표준원가집계")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    NCOLS = 10
    HEADERS = [
        "순번", "제품코드", "제품명",
        "직접재료비합계(원)", "직접노무비합계(원)", "제조간접비합계(원)",
        "제조원가합계(원)", "생산수량(개)", "단위당원가(원)", "비고"
    ]

    style_title_row(ws, 1, NCOLS, "표준원가 집계표 — 원가관리팀 취합")
    ws.row_dimensions[1].height = 32

    style_guide_row(ws, 2, NCOLS,
        "  ※ 작성 안내: 직접재료비/노무비/간접비 합계와 생산수량을 입력하세요. "
        "제조원가합계(=재료비+노무비+간접비)와 단위당원가는 자동 계산됩니다.")

    ws.row_dimensions[3].height = 6

    ws.row_dimensions[4].height = 28
    for col, h in enumerate(HEADERS, 1):
        style_header_cell(ws.cell(row=4, column=col), h)

    # D=직접재료비, E=직접노무비, F=제조간접비, G=제조원가합계(수식), H=생산수량, I=단위당원가(수식)
    samples = [
        (1, "P-001", "제품A",  90000, 132000,  82000, 100),
        (2, "P-002", "제품B",  20000,  90000,  55000,  50),
        (3, "P-003", "제품C",  50000,  60000,  40000,  80),
    ]

    for r, row in enumerate(samples, 5):
        seq, pcode, pname, mat, labor, ovhd, qty = row
        ws.row_dimensions[r].height = 18
        dc(ws.cell(r,1), seq,   align=CENTER)
        dc(ws.cell(r,2), pcode, align=CENTER)
        dc(ws.cell(r,3), pname, align=LEFT)
        dc(ws.cell(r,4), mat,   fmt=NUM_FMT)
        dc(ws.cell(r,5), labor, fmt=NUM_FMT)
        dc(ws.cell(r,6), ovhd,  fmt=NUM_FMT)
        # G: 제조원가합계 수식
        cg = ws.cell(r, 7)
        cg.value=f"=D{r}+E{r}+F{r}"; cg.number_format=NUM_FMT
        cg.fill=FORMULA_FILL; cg.alignment=RIGHT; cg.border=thin_border(); cg.font=NORMAL_FONT
        dc(ws.cell(r,8), qty,   fmt=NUM_FMT)
        # I: 단위당원가 수식
        ci = ws.cell(r, 9)
        ci.value=f"=IF(H{r}>0,G{r}/H{r},0)"; ci.number_format=FLOAT_FMT
        ci.fill=FORMULA_FILL; ci.alignment=RIGHT; ci.border=thin_border(); ci.font=NORMAL_FONT
        dc(ws.cell(r,10), "", align=LEFT)

    # 합계행
    tr = 5 + len(samples)
    ws.row_dimensions[tr].height = 22
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=3)
    c = ws.cell(tr, 1)
    c.value="합  계"; c.fill=TOTAL_FILL; c.font=BOLD_FONT; c.alignment=CENTER; c.border=thin_border()
    for col in range(2, 4):
        ws.cell(tr,col).fill=TOTAL_FILL; ws.cell(tr,col).border=thin_border()

    for col, letter in [(4,"D"),(5,"E"),(6,"F"),(7,"G"),(8,"H")]:
        c = ws.cell(tr, col)
        c.value=f"=SUM({letter}5:{letter}{tr-1})"
        c.number_format=NUM_FMT; c.fill=TOTAL_FILL; c.font=BOLD_FONT; c.alignment=RIGHT; c.border=thin_border()
    ws.cell(tr,9).fill=TOTAL_FILL;  ws.cell(tr,9).border=thin_border()
    ws.cell(tr,10).fill=TOTAL_FILL; ws.cell(tr,10).border=thin_border()

    set_col_widths(ws, [6, 12, 14, 16, 16, 16, 16, 12, 14, 16])


# ══════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    build_bom(wb)
    build_process_list(wb)
    build_process_cost(wb)
    build_cost_summary(wb)

    wb.save(OUTPUT_PATH)
    import os
    size = os.path.getsize(OUTPUT_PATH)
    print(f"[OK] 파일 생성 완료: {OUTPUT_PATH}  ({size:,} bytes)")

if __name__ == "__main__":
    main()
